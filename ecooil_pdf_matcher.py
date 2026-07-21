# -*- coding: utf-8 -*-
"""
Eco-Oil PDF matcher — links every synced unload event (EcoOilUnloadEvent)
to its filed certificate PDF. Read-only on Z:; writes only pdf_path in the
local dev DB + a styled completion-list Excel for the leftovers.

Routing rule (Limor 2026-07-13): the PDF is filed under the BILLED party's
folder (direct customer → their לקוחות folder; transporter's own business →
the transporter's מובילים folder). We therefore score folder-owner matches
billed-first, then customer, then transporter.
"""
import os, io, re, unicodedata
from collections import defaultdict
from datetime import date

os.environ.pop("DATABASE_URL", None)

from src.app import create_app
from src.app.db import db, EcoOilUnloadEvent

YEARS = (2024, 2025, 2026)
TREES = [(r"Z:\Eco_General\מובילים", "transporter"),
         (r"Z:\Eco_General\לקוחות", "customer")]
OUT_XLSX = r"C:\eco_oil_portal\השלמת התאמות - אישורי פריקה.xlsx"

FNAME_RE = re.compile(
    r"^(?:\d{1,3}_)?"                                  # optional leading serial
    r"(?P<stream>[^_]+)"
    r"(?:_(?!\d)[^_]+)?"                               # optional variant segment (e.g. 'שלוח אקו')
    r"_(?P<d>\d{1,2})\.(?P<m>\d{1,2})(?:\.(?P<y>\d{2,4}))?"
    r"(?:_(?P<name>.+))?"                              # customer part optional
    r"\.pdf$", re.I)

# ---------- name normalization ----------
STOP = {"בעמ", "בע", "מ", "חברה", "לישראל", "ישראל", "והשקעות", "בעימ"}

def norm(s):
    if not s:
        return ""
    s = str(s)
    s = s.replace('"', "").replace("'", "").replace("_", " ").replace("-", " ")
    s = re.sub(r"בע\s*מ", " ", s)
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def tokens(s):
    return {t for t in norm(s).split() if t not in STOP}

def name_score(a, b):
    ta, tb = tokens(a), tokens(b)
    if not ta or not tb:
        return 0.0
    inter = ta & tb
    if inter:
        return len(inter) / min(len(ta), len(tb))
    na, nb = norm(a).replace(" ", ""), norm(b).replace(" ", "")
    if na and nb and (na in nb or nb in na):
        return 0.6
    return 0.0

# ---------- legacy billed-name aliases ----------
# The 21/07/2026 canonicalization unified 1,264 ריכוז rows to Limor's official
# spellings, but filing FOLDERS on Z: may still carry the old names. Score
# against the old spellings too so those files keep matching.
LEGACY_BILLED_ALIASES = {
    'עידן שאיבות בע"מ': ['עידן שאיבות'],
    'ר.מ. שאיבות': ['ר.מ שאיבות בע"מ', 'ר.מ שאיבות'],
    'מיקוש שאיבות': ['מיקוש שאיבות בע"מ'],
    'א.מ. שאיבות': ['א.מ שאיבות בע"מ'],
    'ש.מ.ע שאיבות': ['ש.מ.ע שאיבות בע"מ'],
    'ורידיס בע"מ': ['ורידיס'],
    'מאסטר הצפון שירותי ביובית בע"מ': ['מאסטר הצפון-שירותי ביובית בע"מ'],
    'גרין סאן שירותי ביובית בע"מ': ['גרין סאן שירותי ביובית'],
    'טמבור בע"מ - עכו דרום': ['טמבור בע"מ, עכו דרום'],
    'טמבור בע"מ - אתר אסקר': ['טמבור בע"מ, אתר אסקר'],
    'אביר מחזור בע"מ': ['אביר מחזור'],
    'ביובית סאבו בע"מ': ['ביובית סאבו'],
    'אור ברקן בע"מ': ['אור ברקן'],
    'א.ש. תעשיות אקולוגיה בע"מ': ['א.ש תעשיות אקולוגיה בע"מ'],
    'מליסרון, קריון': ['מליסרון - קריון'],
    'עלה - גדות מסופים לכימיקלים בע"מ': ['עלה שירותי אקולוגיה בע"מ',
                                          'עלה שירותי אקולוגיה'],
    'לאור הנדסה ואחזקה תעשייתית בע"מ': ['לאור הנדסה בע"מ', 'לאור הנדסה'],
    'אורות פנינה - תחנת כח חגית מזרח': ['אורות פנינה-תחנת כח חגית מזרח'],
    'א.ע ביובית הצפון בע"מ': ['א.ע. ביובית הצפון בע"מ'],
    'ארדן שנאים (קבוצת ניסקו בע"מ)': ['ארדן שנאים, קבוצת ניסקו בע"מ'],
    'איטונג בע"מ, פרדס חנה': ['איטונג בע"מ - פרדס חנה'],
    'תעשיות בית-אל זכרון יעקב בע"מ, אתר שח"ק': ['תעשיות בית אל - שח"ק',
                                                 'תעשיות בית-אל - שח"ק'],
    'O.P.C אנרגיה בע"מ - תחנת כח': ['O.P.C חדרה בע"מ - תחנת כוח',
                                     'OPC חדרה-תחנת כוח'],
    'ישראל בנקר עיבוד שבבי מדוייק בע"מ': ['ישראל בנקר-עיבוד שבבי מדוייק בע"מ'],
    'מ. מסרי לעסקים בע"מ': ['מ.מסרי לעסקים בע"מ'],
    'נילית בע"מ': ['נילית  בע"מ'],
    'אל חי עיבוד שבבי בע"מ': ['אל-חי עיבוד שבבי בע"מ'],
    'אביר מחזור בע"מ (א.מ.י מערכות משולבות בע"מ)': [
        'אביר מחזור (א.מ.י מערכות משולבות בע"מ)',
        'אביר מחזור, א.מ.י מערכות משולבות בע"מ',
        'אביר מחזור, א.מ.י. מערכות משולבות בע"מ'],
    'לאור הנדסה ואחזקה תעשייתית בע"מ (סולבר מוצרי חלבון בע"מ)': [
        'לאור הנדסה בע"מ (סולבר מוצרי חלבון בע"מ)',
        'לאור הנדסה (סולבר מוצרי חלבון בע"מ)'],
    'אביר מחזור בע"מ (מכשירי ייצור בע"מ זינגפר)': [
        'אביר מחזור (מכשירי ייצור בע"מ זינגפר)',
        'אביר מחזור, מכשירי ייצור בע"מ זינגפר'],
    'אביר מחזור בע"מ (תלעד שבב בע"מ)': [
        'אביר מחזור (תלעד שבב בע"מ)'],
    'ש.צ.פ. מוצרי אריזה מתקדמים בע"מ': ['ש.צ.פ מוצרי אריזה מתקדמים בע"מ'],
    'צ. כץ קבלני שאיבה וקידוחים בע"מ': ['צ. כץ קבלני שאיבה בע"מ'],
    'גדות אחסון ושינוע שותפות מוגבלת בע"מ': ['גדות אחסון ושינוע'],
}

def billed_variants(billed):
    """Current billed name + its legacy spellings (for folder/filename scoring)."""
    if not billed:
        return []
    return [billed] + LEGACY_BILLED_ALIASES.get(billed, [])

BASE_STREAMS = ["מי שטיפה", "מינרלי", "אמולסיה", "מזוט", "חומצה", "בסיס",
                "צמחי", "סניטרי", "רכז שפכים"]
FIX = {"מנרלי": "מינרלי", "מנירלי": "מינרלי", "רכז": "רכז שפכים"}

def base_stream(s):
    if not s:
        return None
    n = norm(s)
    n = FIX.get(n, n)
    for b in BASE_STREAMS:
        if n.startswith(norm(b)) or norm(b) in n:
            return b
    return n or None

def month_num(folder):
    m = re.match(r"(\d{1,2})", folder.strip())
    return int(m.group(1)) if m else None

# ---------- 1. index every filed certificate PDF ----------
# Full walk per owner: certificates live in many layouts (year/month, אישורים/year,
# sub-entity folders per site, loose files under a year folder). Month+day come
# from the FILENAME; the year from the filename if present, else the nearest
# year folder on the path. Sub-folders on the way (e.g. a group's site name)
# join the owner for scoring.
# key: (year, month, day, base_stream) -> list of pdf records
index = defaultdict(list)
n_files = 0
YEAR_NAMES = {str(y) for y in YEARS}
SKIP_PARTS = {"אישורים", "ישן"}

def scan_owner(base, owner):
    global n_files
    for dirpath, dirnames, filenames in os.walk(base):
        rel = os.path.relpath(dirpath, base)
        parts = [] if rel == "." else rel.split(os.sep)
        if any("מלווה" in p for p in parts):
            continue
        year_on_path = next((int(p) for p in parts if p in YEAR_NAMES), None)
        # sub-entity names = path parts that are not year/month/bookkeeping folders
        subs = [p for p in parts
                if p not in YEAR_NAMES and p not in SKIP_PARTS
                and not re.fullmatch(r"\d{1,2}([\./]\d{2,4})?", p)]
        owner_full = " ".join([owner] + subs)
        for f in filenames:
            if not f.lower().endswith(".pdf"):
                continue
            if "מלווה" in f or "שקילה" in f:
                continue
            m = FNAME_RE.match(f)
            if not m:
                continue
            d, mo = int(m["d"]), int(m["m"])
            if not (1 <= d <= 31 and 1 <= mo <= 12):
                continue
            y = None
            if m["y"]:
                yy = int(m["y"])
                y = yy + 2000 if yy < 100 else yy
            if y is None:
                y = year_on_path
            if y is None or y not in YEARS:
                continue
            rec = {"path": os.path.join(dirpath, f), "owner": owner_full,
                   "name": m["name"], "stream": base_stream(m["stream"]),
                   "used": False}
            index[(y, mo, d, rec["stream"])].append(rec)
            n_files += 1

for tree, kind in TREES:
    for owner in os.listdir(tree):
        p = os.path.join(tree, owner)
        if os.path.isdir(p):
            scan_owner(p, owner)

# ---------- 2. match every event ----------
app = create_app()
log = io.StringIO()
with app.app_context():
    events = (EcoOilUnloadEvent.query
              .filter(EcoOilUnloadEvent.event_date.isnot(None))
              .order_by(EcoOilUnloadEvent.event_date, EcoOilUnloadEvent.serial)
              .all())
    log.write(f"events: {len(events)} | indexed PDFs: {n_files}\n")

    def owner_score(rec, ev):
        s = 0.0
        for b in billed_variants(ev.billed_to):
            s = max(s, name_score(rec["owner"], b) * 1.2)  # billed first
        if ev.customer:
            s = max(s, name_score(rec["owner"], ev.customer))
        if ev.transporter:
            s = max(s, name_score(rec["owner"], ev.transporter) * 0.9)
        return s

    matched = 0
    unmatched = []
    for ev in events:
        d = ev.event_date
        bs = base_stream(ev.stream)
        cands = index.get((d.year, d.month, d.day, bs), [])
        best, best_score = None, 0.0
        for rec in cands:
            if rec["used"]:
                continue
            osc = owner_score(rec, ev)
            if osc < 0.5:
                continue
            if rec["name"]:
                nsc = max([name_score(rec["name"], ev.customer or "")] +
                          [name_score(rec["name"], b)
                           for b in billed_variants(ev.billed_to)])
                # weak/internal filename (e.g. 'קוביות', 'בריכה צפונית_160') is fine
                # when the file sits in the billed party's own folder on the right
                # day+stream — that is already strong evidence.
                if nsc < 0.4 and osc < 1.0:
                    continue
            else:
                # no customer in the filename (older style) — trust a strong owner match
                if osc < 0.7:
                    continue
                nsc = 0.0
            score = osc + nsc
            if score > best_score:
                best, best_score = rec, score
        if best:
            best["used"] = True
            ev.pdf_path = best["path"]
            matched += 1
        else:
            unmatched.append(ev)
    db.session.commit()

    # Rows that never expect a certificate leave the gap list (Limor 20/07/2026):
    # 'no_cert_by_design' = closed for good; 'awaiting_declaration' = its own
    # follow-up list (the sanction rows shown in the portal).
    awaiting = [ev for ev in unmatched if ev.doc_status == "awaiting_declaration"]
    closed_by_design = [ev for ev in unmatched if ev.doc_status == "no_cert_by_design"]
    unmatched = [ev for ev in unmatched if not ev.doc_status]

    total = len(events)
    log.write(f"matched: {matched} ({matched*100//total}%)\n")
    log.write(f"unmatched (real gaps): {len(unmatched)}\n")
    log.write(f"awaiting declaration (sanction rows): {len(awaiting)}\n")
    log.write(f"closed by design (no cert ever): {len(closed_by_design)}\n")
    per_year = defaultdict(lambda: [0, 0])
    for ev in events:
        per_year[ev.event_date.year][0] += 1
        if ev.pdf_path:
            per_year[ev.event_date.year][1] += 1
    for y in sorted(per_year):
        t, m = per_year[y]
        log.write(f"  {y}: {m}/{t} ({m*100//t}%)\n")

    # ---------- 3. completion list (styled, RTL, bordered — per the rule) ----------
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "לא הותאמו"
    ws.sheet_view.rightToLeft = True
    headers = ["שנה", "חודש", "תאריך", "לקוח (מקור)", "חיוב", "מוביל", "זרם",
               "כמות (טון)", "קוד", "גיליון מקור", "שורה", "הערות", "טופל"]
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="3E6F69")
    head_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font, c.border = head_fill, head_font, border
        c.alignment = Alignment(horizontal="center")
    for ev in unmatched:
        ws.append([ev.year, ev.month,
                   ev.event_date.strftime("%d/%m/%Y") if ev.event_date else "",
                   ev.customer, ev.billed_to, ev.transporter, ev.stream,
                   ev.declared_tons, ev.code, ev.source_sheet, ev.source_row,
                   ev.notes, ""])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = border
    widths = [7, 7, 12, 30, 26, 26, 14, 11, 12, 12, 8, 24, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # second sheet: the sanction rows — cert withheld until the declaration lands
    ws2 = wb.create_sheet("ממתינות להצהרה")
    ws2.sheet_view.rightToLeft = True
    ws2.append(headers)
    for c in ws2[1]:
        c.fill, c.font, c.border = head_fill, head_font, border
        c.alignment = Alignment(horizontal="center")
    for ev in awaiting:
        ws2.append([ev.year, ev.month,
                    ev.event_date.strftime("%d/%m/%Y") if ev.event_date else "",
                    ev.customer, ev.billed_to, ev.transporter, ev.stream,
                    ev.declared_tons, ev.code, ev.source_sheet, ev.source_row,
                    ev.notes, ""])
    for row in ws2.iter_rows(min_row=2):
        for c in row:
            c.border = border
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    wb.save(OUT_XLSX)
    log.write(f"\ncompletion list -> {OUT_XLSX}\n")

open(r"C:\eco_oil_platform_git\_ecooil_match_result.md", "w", encoding="utf-8").write(log.getvalue())
print(log.getvalue())
