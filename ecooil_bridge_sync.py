# -*- coding: utf-8 -*-
"""
Eco-Oil bridge — LOCAL/dev sync for unload certificates (אישורי פריקה).
Reads the ריכוז workbooks (read-only, safe while open in Excel) and loads one
row per unload event into the portal DB, mapped BY HEADER NAME with aliases
(the sheet format evolved over the years). History scope: rolling 3 years
(2024+ per Limor/Yoav 2026-07-13). Dev version writes to the local SQLite DB;
the production version will run on the office PC and push to the cloud API.
"""
import os, io, re
from datetime import datetime, date, time

# Safety: never let this dev sync touch the production Postgres.
os.environ.pop("DATABASE_URL", None)

import openpyxl
from src.app import create_app
from src.app.db import db, EcoOilUnloadEvent

BASE = r"Z:\Eco_General\ריכוז חודשי"

HEB_MONTHS = {
    "ינואר": 1, "פברואר": 2, "מארס": 3, "מרץ": 3, "אפריל": 4, "מאי": 5,
    "יוני": 6, "יולי": 7, "אוגוסט": 8, "ספטמבר": 9, "אוקטובר": 10,
    "נובמבר": 11, "דצמבר": 12,
}

# field -> (list of header aliases, kind)
SPEC = {
    "serial":        (["מס' תעודה"], "int"),
    "code":          (["קוד_רנדומלי"], "str"),
    "event_date":    (["תאריך"], "date"),
    "vehicle":       (["מספר הרכב"], "str"),
    "transporter":   (["חברת ההובלה"], "str"),
    "customer":      (["לקוח"], "str"),
    "address":       (["כתובת", "מיקום"], "str"),
    "billed_to":     (["חיוב"], "str"),
    "stream":        (["סיווג החומר"], "str"),
    "weight_in":     (["משקל כניסה"], "float"),
    "weight_out":    (["משקל יציאה"], "float"),
    "weight_net":    (["משקל נטו"], "float"),
    "declared_tons": (["משקל מוצהר"], "float"),
    "package_type":  (["סוג אריזה", "סוג האריזה"], "str"),
    "package_count": (["מס' אריזות", "מס' קוביות", "מס' קוביות/ חביות"], "int"),
    "exit_time":     (["שעת יציאה"], "time"),
    "notes":         (["הערות"], "str"),
}

DATE_RE = re.compile(r"\s*(\d{1,2})[/\.](\d{1,2})[/\.](\d{2,4})")

def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None:
        return None
    m = DATE_RE.match(str(v))
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    try:
        return date(y, mo, d)
    except ValueError:
        return None

def to_time(v):
    if isinstance(v, (datetime, time)):
        return v.strftime("%H:%M")
    if v is None:
        return None
    s = str(v).strip()
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    return f"{int(m.group(1)):02d}:{m.group(2)}" if m else (s[:20] or None)

def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def to_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None

def to_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None

COERCE = {"date": to_date, "time": to_time, "int": to_int, "float": to_float, "str": to_str}

# --- Stream normalization (dictionary approved by Limor 2026-07-14) ---
# The RULE: the stream NAME always determines the classification; additions
# (בוצה/פולימר/תשטיפי בטון/לחו"ל...) are internal billing markers. Display
# keeps the original text; filtering uses the canonical stream.
STREAM_VARIANTS = [
    # (canonical, [spellings to search inside the text])
    ("מי שטיפה", ["מי שטיפה", "מי  שטיפה"]),
    ("מינרלי",   ["מינרלי", "מנרלי", "מנירלי"]),
    ("אמולסיה",  ["אמולסיה"]),
    ("סניטרי",   ["סניטרי"]),
    ("בסיס",     ["בסיס"]),
    ("חומצה",    ["חומצה"]),
    ("מזוט",     ["מזוט"]),
    ("צמחי",     ["צמחי"]),
    ("רכז שפכים", ["רכז שפכים"]),
]
# Destruction-service rows (Yoav's accept-and-forward service): no unload
# certificate exists — the customer-facing doc is the signed manifest.
DESTRUCTION_VALUES = {"אריזות מזוהמות", "ג'ריקנים ריקים", "כרומטים", "פילטרים ריקים"}
UNCLASSIFIED_VALUES = {"ממתין למיון", "ללא סיווג", "***", "ללא פירוט"}

def normalize_stream(text):
    if not text:
        return "לא מסווג"
    t = str(text).strip()
    if t in DESTRUCTION_VALUES:
        return "פסולת להשמדה"
    if t in UNCLASSIFIED_VALUES:
        return "לא מסווג"
    # earliest canonical-name occurrence in the text wins
    # (e.g. "מינרלי (מזוט)" → מינרלי, "בוצת אמולסיה" → אמולסיה)
    best, best_pos = None, None
    for canonical, spellings in STREAM_VARIANTS:
        for sp in spellings:
            pos = t.find(sp)
            if pos >= 0 and (best_pos is None or pos < best_pos):
                best, best_pos = canonical, pos
    return best or "לא מסווג"

def sheet_month(title):
    """Extract the month number from a monthly-sheet title, or None to skip."""
    first = title.strip().split()[0].split("_")[0]
    return HEB_MONTHS.get(first)

def load_sheet(ws, year, month, log):
    """Load one monthly sheet. Returns number of rows loaded."""
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return 0
    idx = {}
    for field, (aliases, _) in SPEC.items():
        for a in aliases:
            for i, h in enumerate(header):
                if h is not None and str(h).strip() == a:
                    idx[field] = i
                    break
            if field in idx:
                break
    missing = [f for f in ("event_date", "customer", "stream") if f not in idx]
    if missing:
        log.write(f"    SKIP {ws.title}: missing key headers {missing}\n")
        return 0

    n, empty_streak = 0, 0
    for excel_row, row in enumerate(it, start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            empty_streak += 1
            if empty_streak >= 50:      # junk-formatted sheets report 1M rows
                break
            continue
        empty_streak = 0
        kwargs = {}
        for field, (aliases, kind) in SPEC.items():
            i = idx.get(field)
            v = row[i] if (i is not None and i < len(row)) else None
            kwargs[field] = COERCE[kind](v)
        d = kwargs["event_date"]
        if d is None or not (kwargs["customer"] or kwargs["transporter"]):
            continue
        if d.year != year or d.month != month:
            # tolerate stray rows but keep them under their sheet's month
            pass
        db.session.add(EcoOilUnloadEvent(
            year=year, month=month, source_sheet=ws.title, source_row=excel_row,
            stream_norm=normalize_stream(kwargs.get("stream")),
            **kwargs))
        n += 1
    return n

def load_workbook_months(path, year, log, only_sheet=None, month_override=None):
    if not os.path.exists(path):
        log.write(f"  MISSING FILE: {path}\n")
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    total = 0
    try:
        for ws in wb.worksheets:
            if only_sheet is not None:
                if ws.title.strip() != only_sheet:
                    continue
                month = month_override
            else:
                month = sheet_month(ws.title)
                if month is None:
                    continue
            n = load_sheet(ws, year, month, log)
            total += n
            if n:
                log.write(f"    {ws.title}: {n} rows\n")
    finally:
        wb.close()
    return total

log = io.StringIO()
app = create_app()
with app.app_context():
    EcoOilUnloadEvent.__table__.drop(db.engine, checkfirst=True)
    db.create_all()
    grand = 0

    log.write("2026 (pivot):\n")
    grand += load_workbook_months(rf"{BASE}\2026\ריכוז2026_pivot.xlsx", 2026, log)

    log.write("2025 Sep-Dec (pivot):\n")
    grand += load_workbook_months(rf"{BASE}\2025\ריכוז_pivot.xlsx", 2025, log)

    log.write("2025 Jan-Aug (per-month לרוית files, sheet 'ריכוז מלא'):\n")
    LAVIT = {
        1: "ינואר2025.xlsx", 2: "פברואר 25.xlsx", 3: "מארס 2025.xlsx",
        4: "אפריל 2025.xlsx", 5: "מאי  2025.xlsx", 6: "יוני 2025.xlsx",
        7: "יולי 2025.xlsx", 8: "אוגוסט_2025.xlsx",
    }
    for month, fname in LAVIT.items():
        grand += load_workbook_months(rf"{BASE}\2025\לרוית\{fname}", 2025, log,
                                      only_sheet="ריכוז מלא", month_override=month)

    log.write("2024 (yearly file):\n")
    grand += load_workbook_months(rf"{BASE}\2024\ריכוז 2024.xlsx", 2024, log)

    db.session.commit()

    # summary
    log.write(f"\nTOTAL loaded: {grand}\n\nper year/month:\n")
    from sqlalchemy import func
    q = (db.session.query(EcoOilUnloadEvent.year, EcoOilUnloadEvent.month,
                          func.count(EcoOilUnloadEvent.id))
         .group_by(EcoOilUnloadEvent.year, EcoOilUnloadEvent.month)
         .order_by(EcoOilUnloadEvent.year, EcoOilUnloadEvent.month))
    for y, m, c in q:
        log.write(f"  {y}-{m:02d}: {c}\n")
    q2 = (db.session.query(EcoOilUnloadEvent.stream, func.count(EcoOilUnloadEvent.id))
          .group_by(EcoOilUnloadEvent.stream)
          .order_by(func.count(EcoOilUnloadEvent.id).desc()))
    log.write("\nper stream:\n")
    for s, c in q2:
        log.write(f"  {c:>5}  {s}\n")

open(r"C:\eco_oil_platform_git\_ecooil_bridge_result.md", "w", encoding="utf-8").write(log.getvalue())
print(log.getvalue())
