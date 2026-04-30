# -*- coding: utf-8 -*-
"""Add Materials lookup sheet + column AD (UN number) to טנקו_NEW_DESIGN.xlsx,
and update cleaning cert section 7 (Previous load) to pull the UN number.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC = r'C:/Users/lio2e/Documents/AI_DEV/eco_oil_platform/טנקו_NEW_DESIGN.xlsx'
PRICELIST = r'D:/אקו_דיפו/מחירון יואב.xlsx'

# ===== Step 1: read materials from price list =====
wb_p = openpyxl.load_workbook(PRICELIST, data_only=True)
ws_p = wb_p['גיליון2']
materials = []
for r in range(2, ws_p.max_row + 1):
    name = ws_p.cell(row=r, column=2).value  # B
    un   = ws_p.cell(row=r, column=3).value  # C
    if name and str(name).strip():
        un_str = str(un).strip() if un is not None else ''
        materials.append((str(name).strip(), un_str))
print(f'Read {len(materials)} materials from price list.')

# ===== Step 2: open Tanko file =====
wb = openpyxl.load_workbook(SRC)

# Remove old Materials sheet if exists
if 'Materials' in wb.sheetnames:
    del wb['Materials']

# Create Materials sheet at the end
ws_m = wb.create_sheet('Materials')
ws_m.sheet_view.rightToLeft = False
ws_m.column_dimensions['A'].width = 6
ws_m.column_dimensions['B'].width = 50
ws_m.column_dimensions['C'].width = 14

# Header
ws_m['A1'] = '#'
ws_m['B1'] = 'NAME OF CHEMICAL SUBSTANCE'
ws_m['C1'] = 'UN NUMBER'
hdr_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='333333')
for col in ['A', 'B', 'C']:
    ws_m[f'{col}1'].font = hdr_font
    ws_m[f'{col}1'].fill = hdr_fill
    ws_m[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')

# Populate
for i, (name, un) in enumerate(materials, start=2):
    ws_m.cell(row=i, column=1, value=i - 1)
    ws_m.cell(row=i, column=2, value=name)
    ws_m.cell(row=i, column=3, value=un)

# Hide the sheet from the employee's view (she can unhide via right-click if needed)
ws_m.sheet_state = 'hidden'

# ===== Step 3: Add column AD ("מספר או"מ") to monthly logs =====
MONTHLY = ['ינואר', 'פבואר ', 'מרץ', 'אפריל']
for sname in MONTHLY:
    if sname not in wb.sheetnames:
        continue
    ws_log = wb[sname]
    # Header at row 2 (which is where this file's headers live)
    ws_log.cell(row=2, column=30, value='מספר או"מ')  # AD = 30
    ws_log.column_dimensions['AD'].width = 14
    # Add VLOOKUP formula for every existing data row that has a material in col E
    max_r = ws_log.max_row
    for r in range(3, max_r + 1):
        material_cell = ws_log.cell(row=r, column=5).value  # E
        if material_cell and str(material_cell).strip():
            formula = (f'=IFERROR(VLOOKUP(E{r},Materials!B:C,2,FALSE),"")')
            ws_log.cell(row=r, column=30, value=formula)

# ===== Step 4: Update cleaning cert section 7 (Previous load row) =====
ws_c = wb['תעודת שטיפה']
# Section 7 data row contains 3 columns: Comp (B), UN N° (C), Name (D-J)
# Find the Previous load data row by scanning for the 'Previous load' label
prev_label_row = None
for r in range(5, ws_c.max_row + 1):
    v = ws_c.cell(row=r, column=2).value
    if v == 'Previous load':
        prev_label_row = r
        break
if prev_label_row:
    # data row is prev_label_row + 2 (label, headers, data)
    data_r = prev_label_row + 2
    ws_c.cell(row=data_r, column=2, value=1)  # Comp
    ws_c.cell(row=data_r, column=3, value='=אפריל!AD198')  # UN N°
    ws_c.cell(row=data_r, column=4, value='=אפריל!E198')   # Name
    print(f'Updated cert Previous load at row {data_r}.')

wb.save(SRC)
print('Saved:', SRC)
