# -*- coding: utf-8 -*-
r"""
Eco-Oil bridge — MASAD FEED stage (לימור 10/08/2026; עקרונות אושרו 03/08).

Pulls final-approved producer declarations from the portal cloud and writes
them into the masad workbook on Z: —
  * גיליון "הצהרות":        ALWAYS a new row (the ledger; ends the red-font era).
  * גיליון "ח.פ.-היתר-תוקף": update-in-place by ח.פ. — the stream's expiry+class
    pair only (renewal). Unknown ח.פ. = NEW customer → new row with name/ח.פ./
    permit + the stream pair. Ambiguity (duplicate ח.פ., missing ח.פ., unknown
    stream) is NEVER guessed — reported back as a note; the cloud emails Limor
    once, and the missing half completes automatically on a later cycle.

Safety (the hard-won rules):
  * CreateFileW exclusive probe BEFORE touching the file — if the masad is open
    anywhere (Limor edits it!), skip this cycle entirely.
  * Backup to C:\eco_oil_portal\backups\masad\ before any write (keep last 40).
  * Writing via a PRIVATE invisible Excel instance (DispatchEx only — openpyxl
    would silently strip the masad's data-validation dropdowns), every COM call
    wrapped in the busy-retry helper (the 24/07 "Open.Worksheets" lesson).
  * assert not wb.ReadOnly after open; disk-sentinel verify (openpyxl read-only
    re-read) AFTER save — only verified writes are acked to the cloud.

Usage:
  python ecooil_masad_feed.py                    # production (Z:, real API)
  python ecooil_masad_feed.py --api-base http://127.0.0.1:5000 --masad-path C:\...copy.xlsx
  python ecooil_masad_feed.py --dry-run          # pull + plan, no write, no ack
"""
import argparse
import ctypes
import io
import os
import shutil
import sys
import time
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
load_dotenv(r"C:\eco_oil_platform_git\.env")

import requests

MASAD_PATH = r"Z:\Eco_General\מסד מלא_הצהרות_היתרים_מובילים.xlsx"
BACKUP_DIR = r"C:\eco_oil_portal\backups\masad"
BACKUP_KEEP = 40
DEFAULT_API_BASE = "https://portal.eco-oil.co.il"

LOG_SHEET = "הצהרות"
SUMMARY_SHEET = "ח.פ.-היתר-תוקף"

# stream (english key from the portal) → 1-based column of the expiry in the
# summary sheet; the classification column is always expiry+1.
# Sheet order: F מינרלי H אמולסיה J בסיס L חומצה N מי-שטיפה P מזוט.
STREAM_COL = {
    "mineral": 6,
    "emulsion": 8,
    "base": 10,
    "acid": 12,
    "washwater": 14,
    "gasoil": 16,
}

XL_UP = -4162


# --------------------------------------------------------------- COM helpers
def _com(fn, *args, **kwargs):
    """Busy-retry wrapper on EVERY COM touch (the cert-app 24/07 lesson)."""
    import pywintypes
    last = None
    for _ in range(40):
        try:
            return fn(*args, **kwargs)
        except pywintypes.com_error as exc:  # noqa: PERF203
            last = exc
            hr = exc.args[0] if exc.args else None
            if hr in (-2147417846, -2146777998, -2146827284):  # busy / retry-later
                time.sleep(0.5)
                continue
            raise
    raise last


def _probe_exclusive(path):
    """True if we can open the file exclusively (nobody has it open).
    The r+b trick lies — CreateFileW with no sharing is the truth."""
    GENERIC_READ = 0x80000000
    GENERIC_WRITE = 0x40000000
    OPEN_EXISTING = 3
    h = ctypes.windll.kernel32.CreateFileW(
        path, GENERIC_READ | GENERIC_WRITE, 0, None, OPEN_EXISTING, 0, None)
    if h == -1 or h == 0xFFFFFFFFFFFFFFFF:
        return False
    ctypes.windll.kernel32.CloseHandle(h)
    return True


def _digits(s):
    """ספרות בלבד. תאי מספר חוזרים מאקסל כ-float (512181017.0) — מנקים
    את ה-.0 לפני ההשוואה, אחרת שום ח.פ. לא מתאים לעולם."""
    if isinstance(s, float) and s.is_integer():
        s = int(s)
    return "".join(ch for ch in str(s or "") if ch.isdigit())


def _norm(s):
    s = str(s or "")
    for ch in ('"', "'", "-", "_", ".", ","):
        s = s.replace(ch, " ")
    s = s.replace("בע מ", " ").replace("בעמ", " ")
    return " ".join(s.split())


def _site_keys(d):
    """הצורות שבהן שורת אתר עשויה להיכתב בגיליון, מנורמלות — שתיהן בשוויון
    מלא: שם העסק בלבד, ושם העסק יחד עם הכתובת (שהיא האתר, לימור 17/08)."""
    name = d.get("producer_name") or ""
    addr = d.get("address") or ""
    keys = {_norm(name)}
    if addr.strip():
        keys.add(_norm(f"{name} {addr}"))
    return {k for k in keys if k}


def _match_by_site_words(cand, d):
    """שורת האתר לפי המילים שמבדילות בין השורות בלבד — או None."""
    token_sets = [set(_norm(r[1]).split()) for r in cand]
    shared = set.intersection(*token_sets) if token_sets else set()
    haystack = set(_norm((d.get("producer_name") or "") + " "
                         + (d.get("address") or "")).split())
    hits = []
    for r, toks in zip(cand, token_sets):
        site = toks - shared - {"אתר"}
        if site and site <= haystack:
            hits.append(r[0])
    return hits[0] if len(hits) == 1 else None


def resolve_summary_row(rows, d):
    """איתור שורת האתר הנכונה בגיליון ח.פ.-היתר-תוקף — לפי הבנת לימור 10/08:
    ח.פ. אחד לכל החברה (כמו מספר רישוי); היתר רעלים מזהה אתר (אבל לא תמיד —
    אצל K.L.A כל האתרים חולקים היתר); בלי מבחין — שם האתר מכריע.

    rows = [(row_idx, name, permit, hp)] מכל הגיליון.
    מחזיר (row_idx, None) | (None, note) | ('NEW', None) — בלי ניחושים."""
    biz = _digits(d.get("business_id"))
    by_hp = [r for r in rows if _digits(r[3]) and _digits(r[3]) == biz]

    if len(by_hp) == 1:
        return by_hp[0][0], None
    if not by_hp:
        # לפני פתיחת שורה חדשה: אולי החברה קיימת עם תא ח.פ. ריק/שונה — לא משכפלים.
        # משווים גם מול שם הכרטיס וצורות הכתיב שלו (19/08, גדות פי גלילות:
        # בהצהרה נכתב "גדות פי גלילות" והשורה בגיליון נשאה את השם המלא —
        # רשת-הביטחון החטיאה ונפתחה שורה כפולה).
        name_keys = _site_keys(d) | _account_forms(d)
        name_hits = [r for r in rows if _norm(r[1]) in name_keys]
        if name_hits:
            return None, (
                f'בגיליון "{SUMMARY_SHEET}", שורה {name_hits[0][0]}: קיימת שורה '
                "באותו שם אך הח.פ. בה ריק או שונה. השלימי שם את הח.פ., "
                "או עדכני ידנית את תאריך התוקף באותה שורה.")
        return "NEW", None

    # כמה אתרים: קודם היתר רעלים (אם יש בהצהרה), אחר כך שם האתר
    cand = by_hp
    p = _digits(d.get("permit_number"))
    if p:
        by_permit = [r for r in cand if _digits(r[2]) == p]
        if len(by_permit) == 1:
            return by_permit[0][0], None
        if by_permit:
            cand = by_permit
    # שם האתר יושב בגיליון בתוך תא הלקוח ("גלבוע ... בע"מ, אתר ספיר"), ואילו
    # בהצהרה הוא בשדה נפרד — "כתובת העסק / מפעל". לכן משווים גם את צירוף
    # שם העסק + הכתובת, ולא רק את השם (לימור 18/08, מקרה מגן שאול/ספיר).
    # עדיין שוויון מלא ולא "מכיל" — כדי ש"אתר ספיר" לא ייתפס על "אתר ספיר 2".
    keys = _site_keys(d)
    by_name = [r for r in cand if _norm(r[1]) in keys]
    if len(by_name) == 1:
        return by_name[0][0], None

    # השוואת מילות-האתר בלבד (לימור 18/08, מקרה גלבוע/מגן שאול): מסירים
    # מכל שורה את המילים המשותפות לכל שורות אותה חברה — מה שנשאר מבדיל בין
    # האתרים ("מגן שאול" מול "ספיר") — ובודקים אילו מהן מופיעות בשם+כתובת
    # שבהצהרה. עוקף גם כתובת רחוב מלאה ("תאנה 2, מגן שאול") וגם הבדלי כתיב
    # בשם החברה ("מכשור" מול "מיכשור"), כי שם החברה כלל לא נכנס להשוואה.
    # עדיין בלי ניחוש: חייבת להיות בדיוק שורה אחת מתאימה.
    site_row = _match_by_site_words(cand, d)
    if site_row is not None:
        return site_row, None
    sites = "; ".join(f'שורה {r[0]} — "{r[1]}"' for r in cand)
    return None, (
        f'בגיליון "{SUMMARY_SHEET}": לחברה יש כמה שורות-אתרים ולא ניתן לקבוע '
        f"לאיזה אתר שייכת ההצהרה. השורות: {sites}. "
        "עדכני ידנית את תאריך התוקף בשורה הנכונה. "
        "כדי שזה יזוהה אוטומטית בפעם הבאה — מלאי מספר היתר רעלים בשורות, "
        "או ודאי ששם האתר בתא הלקוח זהה לכתובת שנרשמה בהצהרה.")


def _fmt_date(iso):
    """ISO → dd/mm/yyyy text (the log sheet's issue-date style)."""
    if not iso:
        return ""
    return datetime.fromisoformat(iso).strftime("%d/%m/%Y")


def _parse_dt(iso):
    return datetime.fromisoformat(iso) if iso else None


def _xl_serial(dt):
    """datetime → מספר סידורי של אקסל. כתיבת datetime ישירות דרך COM עוברת
    המרת אזור-זמן ומזיזה את התאריך 3 שעות אחורה (31/08 הפך 30/08 21:00 בבדיקה);
    המספר הסידורי דטרמיניסטי ואקסל מציג אותו לפי פורמט העמודה."""
    return (dt - datetime(1899, 12, 30)).total_seconds() / 86400.0


# --------------------------------------------------------------- planning
def _account_forms(d):
    """כל צורות הכתיב של החשבון, מנורמלות: שם הכרטיס + צורות הכתיב שבו."""
    forms = [d.get("account_name")] + list(d.get("account_aliases") or [])
    return {_norm(x) for x in forms if x}


def plan_client_type(d):
    """עמודת "סוג לקוח" לשורת לקוח חדשה — בלי ניחושים:
    ישיר כשהחשבון הוא היצרן עצמו — גם כשההצהרה נכתבה באחת מצורות הכתיב
    של הכרטיס (לקח 19/08, גדות פי גלילות + גרין סטרים: שם מקוצר/מלא בהצהרה
    נראה כמו "מוביל שמילא ללקוחו" ושם החשבון נרשם בעמודת סוג-לקוח);
    שם המוביל כשהמוביל מילא עבור לקוחו;
    ריק + הערה כשהיצרן העקיף הגיש בעצמו (המוביל לא ידוע למערכת)."""
    if d.get("account_type") == "indirect":
        return "", (f'בגיליון "{SUMMARY_SHEET}": נוספה שורת לקוח חדשה — השלימי '
                    'את עמודת "סוג לקוח" (שם המוביל).')
    if _norm(d.get("producer_name")) in _account_forms(d):
        return "ישיר", None
    return d.get("account_name") or "", None


# --------------------------------------------------------------- main work
def run(api_base, masad_path, dry_run=False):
    token = os.environ.get("ECOOIL_BRIDGE_TOKEN")
    if not token:
        print("ERROR: ECOOIL_BRIDGE_TOKEN missing from .env")
        return 1
    headers = {"Authorization": "Bearer " + token}

    r = requests.get(api_base + "/bridge/ecooil/masad-feed", headers=headers, timeout=60)
    if r.status_code != 200:
        print(f"ERROR: masad-feed pull failed {r.status_code}: {r.text[:200]}")
        return 1
    decls = r.json().get("declarations") or []
    if not decls:
        print("masad-feed: nothing pending")
        return 0
    print(f"masad-feed: {len(decls)} declaration(s) pending")

    if dry_run:
        for d in decls:
            print(f"  would write #{d['id']} {d['producer_name']} ({d['material_name']})"
                  f" log={d['log_pending']} summary={d['summary_pending']}")
        return 0

    if not os.path.exists(masad_path):
        print(f"ERROR: masad not found at {masad_path}")
        return 1

    # --- Excel isolation: never touch an open file ---
    if not _probe_exclusive(masad_path):
        print("masad is OPEN somewhere (probably Limor's Excel) — skipping this cycle")
        return 0

    # --- backup to C: before any write ---
    os.makedirs(BACKUP_DIR, exist_ok=True)
    backup = os.path.join(
        BACKUP_DIR, f"מסד_{datetime.now():%Y-%m-%d_%H%M%S}.xlsx")
    shutil.copy2(masad_path, backup)
    olds = sorted(f for f in os.listdir(BACKUP_DIR) if f.endswith(".xlsx"))
    for f in olds[:-BACKUP_KEEP]:
        os.remove(os.path.join(BACKUP_DIR, f))
    print(f"backup: {backup}")

    import pythoncom
    from win32com.client import DispatchEx

    results = {d["id"]: {"id": d["id"], "log_done": not d["log_pending"],
                         "summary_done": not d["summary_pending"], "note": None}
               for d in decls}
    sentinels = []  # (sheet, row, col, expected, decl_id, kind)

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = _com(excel.Workbooks.Open, masad_path)
        if wb.ReadOnly:
            print("workbook opened READ-ONLY — someone holds it; skipping cycle")
            _com(wb.Close, False)
            return 0

        ws_log = _com(wb.Worksheets, LOG_SHEET)
        ws_sum = _com(wb.Worksheets, SUMMARY_SHEET)

        # עמודת AB "נקלט על ידי" (לימור 25/08): מוודאים שהכותרת קיימת פעם אחת.
        # "מטעם" (AA) חוזרת לתפקידה — מי הביא את הלקוח; סימון הפורטל עבר ל-AB.
        if not str(_com(getattr, ws_log.Cells(1, 28), "Value") or "").strip():
            _com(setattr, ws_log.Cells(1, 28), "Value", "נקלט על ידי")

        # אינדקס הגיליון המסכם ל"מטעם" (הכרעת לימור 25/08, מקרה אסתילון/אב-גל):
        # המקור המוסמך לשיוך הוא עמודת "סוג לקוח" שהיא מתחזקת בגיליון המסכם —
        # אם החברה כבר שם והעמודה מלאה, מעתיקים אותה; רק בהיעדרה נגזר השיוך
        # מהחשבון המגיש (plan_client_type). נטען פעם אחת בתחילת הסבב.
        _sum_last = max(
            _com(ws_sum.Cells(ws_sum.Rows.Count, 2).End, XL_UP).Row,
            _com(ws_sum.Cells(ws_sum.Rows.Count, 5).End, XL_UP).Row)
        _sum_grid = _com(ws_sum.Range,
                         ws_sum.Cells(2, 1), ws_sum.Cells(_sum_last, 5)).Value
        _sum_grid = _sum_grid if isinstance(_sum_grid, tuple) else (_sum_grid,)
        _sum_rows_ix = [(i + 2, r[1], r[2], r[4]) for i, r in enumerate(_sum_grid)]
        _sum_col_a = {i + 2: str(r[0]).strip() if r[0] else ""
                      for i, r in enumerate(_sum_grid)}

        def _curated_referrer(d):
            """עמודת "סוג לקוח" מהגיליון המסכם עבור היצרן — או ריק."""
            try:
                target, _n = resolve_summary_row(_sum_rows_ix, d)
            except Exception:
                return ""
            return _sum_col_a.get(target, "") if isinstance(target, int) else ""

        for d in decls:
            res = results[d["id"]]
            note_parts = []
            try:
                # ---- half 1: the ledger — always a new row ----
                if d["log_pending"]:
                    last = _com(ws_log.Cells(ws_log.Rows.Count, 2).End, XL_UP).Row
                    row = last + 1
                    vals = {
                        2: d.get("producer_name"), 3: d.get("address"),
                        4: d.get("business_id"), 5: d.get("permit_number"),
                        6: d.get("producer_size"), 7: d.get("material_name"),
                        8: d.get("waste_stream_number"), 9: d.get("production_facility"),
                        10: d.get("y_code"), 11: d.get("annex8"), 12: d.get("h_code"),
                        13: d.get("un_group"), 14: d.get("catalog"),
                        15: d.get("treatment_type"), 16: d.get("r_code"),
                        17: d.get("d_code"), 18: d.get("quantity"),
                        19: d.get("packaging"), 20: d.get("characteristic"),
                        21: d.get("pollutant_type"), 22: d.get("concentration_range"),
                        23: d.get("addressed_to"), 24: d.get("producer_email"),
                        # גרש מוביל = טקסט מאולץ; בלעדיו אקסל הופך "10/08/2026"
                        # לתאריך בפרשנות אמריקאית (8 באוקטובר!)
                        25: "'" + _fmt_date(d.get("valid_from")),
                    }
                    # "מטעם" (AA, לימור 25/08): קודם מה שרשום בגיליון המסכם
                    # (המקור המוסמך לשיוך); רק בהיעדרו — ישיר / שם המוביל לפי
                    # החשבון המגיש; ריק+הערה לעקיף שהגיש בעצמו ואינו במסכם.
                    # "פורטל" עבר ל-AB "נקלט על ידי".
                    aa_val = _curated_referrer(d)
                    if not aa_val:
                        aa_val, _aa_note = plan_client_type(d)
                    if aa_val:
                        vals[27] = aa_val
                    vals[28] = "פורטל"
                    _com(setattr, ws_log.Cells(row, 1), "Formula", "=ROW()-1")
                    for col, v in vals.items():
                        if v not in (None, "", "'"):
                            _com(setattr, ws_log.Cells(row, col), "Value", v)
                    vu = _parse_dt(d.get("valid_until"))
                    if vu:
                        _com(setattr, ws_log.Cells(row, 26), "Value2", _xl_serial(vu))
                    sentinels.append((LOG_SHEET, row, 2, str(d.get("producer_name") or ""),
                                      d["id"], "log"))
                    if not aa_val:
                        note_parts.append(
                            f'בגיליון "{LOG_SHEET}" שורה {row}: השלימי את '
                            'עמודת "מטעם" (שם המוביל של הלקוח העקיף)')
                    print(f"  #{d['id']} log row {row}: {d['producer_name']} / {d['material_name']}")

                # ---- half 2: the validity summary — update by ח.פ. ----
                if d["summary_pending"]:
                    stream_col = STREAM_COL.get(d.get("material_classification"))
                    biz = _digits(d.get("business_id"))
                    if stream_col is None:
                        note_parts.append(
                            f'בגיליון "{SUMMARY_SHEET}": הזרם '
                            f"'{d.get('material_classification')}' אינו אחד מזרמי "
                            "העמודות בגיליון, ולכן לא ידעתי באיזו עמודה לרשום. "
                            "עדכני ידנית את תאריך התוקף בעמודת הזרם המתאימה.")
                    elif not biz:
                        note_parts.append(
                            f'בגיליון "{SUMMARY_SHEET}": אין ח.פ. בהצהרה ולכן לא '
                            "ניתן לאתר את שורת החברה. עדכני ידנית את תאריך התוקף.")
                    else:
                        last_sum = max(
                            _com(ws_sum.Cells(ws_sum.Rows.Count, 2).End, XL_UP).Row,
                            _com(ws_sum.Cells(ws_sum.Rows.Count, 5).End, XL_UP).Row)
                        grid = _com(ws_sum.Range,
                                    ws_sum.Cells(2, 1), ws_sum.Cells(last_sum, 5)).Value
                        grid = grid if isinstance(grid, tuple) else (grid,)
                        # (row_idx, name, permit, hp) — הבסיס לזיהוי אתר
                        rows_ix = [(i + 2, r[1], r[2], r[4]) for i, r in enumerate(grid)]
                        # סוף הנתונים האמיתי — מהערכים עצמם, לא מ-End(xlUp):
                        # הגיליון מעוצב כטבלת-אקסל, ו-End(xlUp) עוצר בגבול
                        # הטבלה גם כששורותיה האחרונות ריקות (19/08: הטבלה
                        # הגיעה עד 553 → גרין סטרים נכתבה ב-554 וגדות פי
                        # גלילות ב-555, במרחק 76 שורות מסוף הנתונים).
                        data_last = max(
                            (i + 2 for i, r in enumerate(grid)
                             if (r[1] is not None and str(r[1]).strip())
                             or (r[4] is not None and str(r[4]).strip())),
                            default=1)
                        target, note = resolve_summary_row(rows_ix, d)
                        vu = _parse_dt(d.get("valid_until"))
                        wnum = d.get("waste_stream_number") or ""
                        if note:
                            note_parts.append(note)
                        elif target == "NEW":
                            ctype, ctype_note = plan_client_type(d)
                            row = data_last + 1
                            newvals = {2: d.get("producer_name"),
                                       3: d.get("permit_number"),
                                       5: d.get("business_id"),
                                       stream_col + 1: wnum}
                            if ctype:
                                newvals[1] = ctype
                            for col, v in newvals.items():
                                if v not in (None, ""):
                                    _com(setattr, ws_sum.Cells(row, col), "Value", v)
                            if vu:
                                _com(setattr, ws_sum.Cells(row, stream_col), "Value2", _xl_serial(vu))
                            sentinels.append((SUMMARY_SHEET, row, 5,
                                              str(d.get("business_id") or ""), d["id"], "summary"))
                            if ctype_note:
                                note_parts.append(ctype_note)
                            print(f"  #{d['id']} summary NEW customer row {row}: {d['producer_name']}")
                        else:
                            row = target
                            if vu:
                                _com(setattr, ws_sum.Cells(row, stream_col), "Value2", _xl_serial(vu))
                            _com(setattr, ws_sum.Cells(row, stream_col + 1), "Value", wnum)
                            sentinels.append((SUMMARY_SHEET, row, stream_col + 1, wnum,
                                              d["id"], "summary"))
                            print(f"  #{d['id']} summary row {row}: {d['material_classification']}"
                                  + (f" → {vu:%d/%m/%Y}" if vu else ""))
            except Exception as exc:
                note_parts.append(f"שגיאת כתיבה בקובץ המסד: {exc}")
                print(f"  #{d['id']} ERROR: {exc}")
            if note_parts:
                res["note"] = "; ".join(note_parts)

        _com(wb.Save)
        _com(wb.Close, True)
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    # --- disk-sentinel verify: trust only what the saved file actually shows ---
    from openpyxl import load_workbook
    verified = {}
    try:
        vb = load_workbook(masad_path, read_only=True, data_only=False)
        for sheet, row, col, expected, decl_id, kind in sentinels:
            actual = vb[sheet].cell(row=row, column=col).value
            ok = str(actual or "").strip() == str(expected or "").strip()
            verified[(decl_id, kind)] = ok
            if not ok:
                print(f"  SENTINEL MISMATCH {sheet} r{row}c{col}: expected '{expected}' got '{actual}'")
        vb.close()
    except Exception as exc:
        print(f"ERROR: sentinel verify failed to open masad: {exc}")
        return 1

    for sheet, row, col, expected, decl_id, kind in sentinels:
        if verified.get((decl_id, kind)):
            results[decl_id]["log_done" if kind == "log" else "summary_done"] = True
        else:
            prev = results[decl_id].get("note")
            msg = "אימות הכתיבה נכשל — ראי לוג גשר"
            results[decl_id]["note"] = f"{prev}; {msg}" if prev else msg

    # --- ack to the cloud (alerts email fires there on new notes) ---
    # masad_path נשלח כדי שהמייל יוכל לומר במפורש לאיזה קובץ ללכת (לימור 18/08)
    r = requests.post(api_base + "/bridge/ecooil/masad-feed/ack", headers=headers,
                      json={"results": list(results.values()),
                            "masad_path": masad_path}, timeout=60)
    if r.status_code != 200:
        print(f"ERROR: ack failed {r.status_code}: {r.text[:200]}")
        return 1
    body = r.json()
    print(f"ack: updated={body.get('updated')} alert_emailed={body.get('alert_emailed')}")
    return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--api-base", default=DEFAULT_API_BASE)
    ap.add_argument("--masad-path", default=MASAD_PATH)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    sys.exit(run(args.api_base, args.masad_path, dry_run=args.dry_run))


if __name__ == "__main__":
    main()
