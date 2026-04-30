# -*- coding: utf-8 -*-
"""Build new-design release certificate sheet inside טנקו_NEW_DESIGN.xlsx."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image as PILImage

SRC = r'C:/Users/lio2e/Documents/AI_DEV/eco_oil_platform/טנקו_NEW_DESIGN.xlsx'
ASSETS = r'C:/Users/lio2e/Documents/AI_DEV/eco_oil_platform/website/images/cert'
EMU = 9525

wb = openpyxl.load_workbook(SRC)

# Remove old release cert sheets
for n in ['תעודת יציאה ', 'תעודת שחרור', 'תעודת יציאה']:
    if n in wb.sheetnames:
        del wb[n]

ws = wb.create_sheet('תעודת שחרור')
ws.sheet_view.rightToLeft = True  # Hebrew RTL

# ===== Column widths =====
COL_W = {'A': 3.5, 'B': 10, 'C': 12, 'D': 12, 'E': 3.5,
         'F': 3.5, 'G': 10, 'H': 12, 'I': 14, 'J': 3.5}
for col, w in COL_W.items():
    ws.column_dimensions[col].width = w

# ===== Styles =====
thin = Side(border_style='thin', color='000000')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

font_title    = Font(name='Arial', size=14, bold=True, color='000000')
font_certno   = Font(name='Arial', size=12, bold=True, color='000000')
font_sub      = Font(name='Arial', size=8,  italic=True, color='000000')
font_url      = Font(name='Arial', size=8,  color='666666')
font_idx      = Font(name='Arial', size=11, bold=True)
font_label    = Font(name='Arial', size=8,  bold=True, color='666666')
font_value    = Font(name='Arial', size=10)
font_tank_lbl = Font(name='Arial', size=10, bold=True, color='666666')
font_tank_val = Font(name='Arial', size=18, bold=True)
font_greet    = Font(name='Arial', size=11, bold=True)
font_foot     = Font(name='Arial', size=8, color='444444')
font_cap      = Font(name='Arial', size=9, bold=True)

fill_black  = PatternFill('solid', fgColor='000000')
fill_lightg = PatternFill('solid', fgColor='F0F0F0')

center     = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_top  = Alignment(horizontal='right',  vertical='top',    wrap_text=True)
right_mid  = Alignment(horizontal='right',  vertical='center', wrap_text=True)


def merge(rng): ws.merge_cells(rng)


def set_cell(addr, value=None, font=None, fill=None, align=None, border=True):
    cell = ws[addr]
    if value is not None: cell.value = value
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border_all
    return cell


def setm(rng, value, font=None, fill=None, align=None, border=False):
    merge(rng)
    cell = ws[rng.split(':')[0]]
    cell.value = value
    if font:  cell.font = font
    if fill:  cell.fill = fill
    if align: cell.alignment = align
    if border:
        for r in ws[rng]:
            for c in (r if hasattr(r, '__iter__') else [r]):
                c.border = border_all


def add_centered_image(img_path, anchor_col_idx, anchor_row_idx,
                       max_width_px, max_height_px,
                       area_width_px, area_height_px):
    pil = PILImage.open(img_path)
    orig_w, orig_h = pil.size
    ratio = orig_w / orig_h
    if max_width_px / max_height_px > ratio:
        h = max_height_px
        w = int(h * ratio)
    else:
        w = max_width_px
        h = int(w / ratio)
    off_x = max(0, (area_width_px - w) // 2)
    off_y = max(0, (area_height_px - h) // 2)
    img = XLImage(img_path)
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=anchor_col_idx, row=anchor_row_idx,
                           colOff=off_x * EMU, rowOff=off_y * EMU),
        ext=XDRPositiveSize2D(cx=w * EMU, cy=h * EMU),
    )
    ws._images.append(img)


# ============ HEADER (rows 1-4) ============
# In RTL view: column A renders on visual RIGHT, column J on visual LEFT.
# So: Logo in A:B (visual right), Title in C:H (middle), QR in I:J (visual left).
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 22
ws.row_dimensions[3].height = 16
ws.row_dimensions[4].height = 14

# Logo cell (visual right)
merge('A1:B4')
ws['A1'].fill = PatternFill('solid', fgColor='FFFFFF')
ws['A1'].border = border_all

# Title bars
setm('C1:H1', 'תעודת שחרור לטנק (איזוטנק)', font_title, None, center, border=True)
setm('C2:H2', 'ECO-REL-2026-0001', font_certno, None, center, border=True)
setm('C3:H3', 'תעודה רשמית — אקו אויל חץ וירומטל בע"מ', font_sub, None, center, border=False)
setm('C4:H4', 'www.eco-oil.co.il', font_url, None, center, border=False)

# QR cell (visual left)
merge('I1:J4')
ws['I1'].fill = PatternFill('solid', fgColor='FFFFFF')
ws['I1'].alignment = center
ws['I1'].border = border_all

# ============ TANK NUMBER BLOCK (row 5-6) ============
ws.row_dimensions[5].height = 14
ws.row_dimensions[6].height = 36

set_cell('A5', value='מספר טנק (איזוטנק):', font=font_tank_lbl, fill=fill_lightg, align=center)
merge('A5:J5')

set_cell('A6', value='=אפריל!A198', font=font_tank_val, align=center)
merge('A6:J6')

# ============ FORM TABLE ============
row = 7


def section(idx, label, value, value_height=22):
    global row
    ws.row_dimensions[row].height = value_height
    set_cell(f'A{row}', value=str(idx), font=font_idx, fill=fill_lightg, align=center)
    set_cell(f'B{row}', value=label, font=font_label, align=right_mid)
    merge(f'B{row}:C{row}')
    set_cell(f'D{row}', value=value, font=font_value, align=right_mid)
    merge(f'D{row}:J{row}')
    row += 1


# 1 - תחנת שחרור (release station) - HARDCODED
station_text = ('אקו אויל חץ וירומטל בע"מ  |  '
                'רחוב המסילה 1, נשר  |  '
                'ת.ד. 116, נשר 3660101  |  '
                'טל\': +972-54-323-2617  |  '
                'shtifot@eco-oil.co.il')
section(1, 'תחנת שחרור', station_text, value_height=44)

# 2 - חברת ההובלה (carrier) - formula from log col C
section(2, 'חברת ההובלה', '=אפריל!C198', value_height=22)

# 3 - צפי יציאה (expected departure) - formula from log col B
section(3, 'צפי יציאה', '=IFERROR(TEXT(אפריל!B198,"dd/mm/yyyy"),"")', value_height=22)

# 4 - באחסון מתאריך (in storage from date) - formula from log col D
section(4, 'באחסון מתאריך', '=IFERROR(TEXT(אפריל!D198,"dd/mm/yyyy"),"")', value_height=22)

# 5 - יעד (destination) - manual
section(5, 'יעד', '', value_height=22)

# 6 - הערות (comments) - formula from log col H
section(6, 'הערות', '=אפריל!H198', value_height=36)

# 7 - חתימה וחותמת (signature + stamp)
ws.row_dimensions[row].height = 14
set_cell(f'A{row}', value='7', font=font_idx, fill=fill_lightg, align=center)
merge(f'A{row}:A{row+4}')
set_cell(f'B{row}', value='חתימה וחותמת', font=font_label, align=right_mid)
merge(f'B{row}:J{row}')
row += 1

sig_row = row
ws.row_dimensions[row].height = 50
ws.row_dimensions[row + 1].height = 50

# In RTL, "right" is column A. We want signature on the right (CEO sig is the "primary"),
# stamp on the left.
# Signature merged columns A:E (visual right side)
set_cell(f'B{row}', value='', align=center)
merge(f'B{row}:E{row+1}')
# Stamp merged columns F:J (visual left side)
set_cell(f'F{row}', value='', align=center)
merge(f'F{row}:J{row+1}')
row += 2

# Divider row
ws.row_dimensions[row].height = 4
set_cell(f'B{row}', value='')
merge(f'B{row}:E{row}')
set_cell(f'F{row}', value='')
merge(f'F{row}:J{row}')
row += 1

# Captions
ws.row_dimensions[row].height = 18
set_cell(f'B{row}', value='חתימת מנכ"ל – אקו אויל', font=font_cap, align=center)
merge(f'B{row}:E{row}')
set_cell(f'F{row}', value='חותמת החברה', font=font_cap, align=center)
merge(f'F{row}:J{row}')
row += 1

# Greeting row
ws.row_dimensions[row].height = 22
set_cell(f'A{row}', value='בברכה,   אקו אויל חץ וירומטל בע"מ', font=font_greet, align=center)
merge(f'A{row}:J{row}')
row += 1

# Footer image row
ws.row_dimensions[row].height = 130
set_cell(f'A{row}', value='', align=center)
merge(f'A{row}:J{row}')
foot_row = row
row += 1

# Print info
ws.row_dimensions[row].height = 14
set_cell(f'A{row}', value='Printed with status SIGNED by Eco Oil — www.eco-oil.co.il',
         font=font_foot, align=center)
merge(f'A{row}:J{row}')
last_row = row

# ===== Insert images =====
# Logo on visual RIGHT (column A in RTL) — anchor_col_idx=0
add_centered_image(f'{ASSETS}/eco-oil-logo.png',
                   anchor_col_idx=0, anchor_row_idx=0,
                   max_width_px=80, max_height_px=80,
                   area_width_px=100, area_height_px=88)

# QR on visual LEFT (column I in RTL) — anchor_col_idx=8
add_centered_image(f'{ASSETS}/qr-eco-oil.png',
                   anchor_col_idx=8, anchor_row_idx=0,
                   max_width_px=78, max_height_px=78,
                   area_width_px=125, area_height_px=88)

# Signature in B..E area (anchor B = col 1)
add_centered_image(f'{ASSETS}/signature-yoav.png',
                   anchor_col_idx=1, anchor_row_idx=sig_row - 1,
                   max_width_px=120, max_height_px=92,
                   area_width_px=260, area_height_px=100)

# Stamp in F..J area (anchor F = col 5)
add_centered_image(f'{ASSETS}/stamp.png',
                   anchor_col_idx=5, anchor_row_idx=sig_row - 1,
                   max_width_px=240, max_height_px=92,
                   area_width_px=300, area_height_px=100)

# Footer
add_centered_image(f'{ASSETS}/footer-contact.png',
                   anchor_col_idx=0, anchor_row_idx=foot_row - 1,
                   max_width_px=595, max_height_px=128,
                   area_width_px=600, area_height_px=130)

# ===== Page setup =====
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.4
ws.print_area = f'A1:J{last_row}'
ws.print_options.horizontalCentered = True

# Reorder sheets
target = ['ינואר', 'פבואר ', 'מרץ', 'אפריל', 'תעודת שטיפה', 'תעודת שחרור']
order_idx = {n: i for i, n in enumerate(target)}
wb._sheets.sort(key=lambda s: order_idx.get(s.title, 99))

wb.save(SRC)
print('Saved release cert. Last row:', last_row)
