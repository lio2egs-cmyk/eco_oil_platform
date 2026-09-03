# -*- coding: utf-8 -*-
r"""
Eco-Oil bridge — VALIDITY SNAPSHOT push (לימור 03/09/2026).

קורא (קריאה בלבד!) את גיליון "ח.פ.-היתר-תוקף" מהמסד ודוחף תמונת-מצב מלאה
לענן: לכל שורה — עמודת "סוג לקוח" (המוביל המביא), שם הלקוח, ח.פ., ותאריך
התוקף של כל זרם. הפורטל גוזר מזה את טבלת "מצב ההצהרות של הלקוחות שלכם"
בעמוד הבית של כל מוביל — הגיליון של לימור הוא מקור האמת, לא הצהרות הענן.

קריאה דרך openpyxl read-only — לא נוגע בקובץ, עובד גם כשהמסד פתוח אצל לימור.

Usage:
  python ecooil_validity_push.py
  python ecooil_validity_push.py --api-base http://127.0.0.1:5000 --masad-path C:\...copy.xlsx
  python ecooil_validity_push.py --dry-run
"""
import argparse
import io
import json
import re
import sys
from datetime import date, datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
load_dotenv(r"C:\eco_oil_platform_git\.env")

import os
import requests

MASAD_PATH = r"Z:\Eco_General\מסד מלא_הצהרות_היתרים_מובילים.xlsx"
DEFAULT_API_BASE = "https://portal.eco-oil.co.il"
SUMMARY_SHEET = "ח.פ.-היתר-תוקף"

# עמודות הגיליון: A סוג לקוח, B שם, C היתר, E ח.פ.; זוגות תוקף+סיווג לפי
# ecooil_masad_feed.STREAM_COL — התוקף בעמודה, הסיווג בעמודה שאחריה.
STREAM_COLS = [(6, "מינרלי"), (8, "אמולסיה"), (10, "בסיס"),
               (12, "חומצה"), (14, "מי שטיפה"), (16, "מזוט")]

_DATE_RE = re.compile(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})$")


def _fmt(v):
    """ערך תא תוקף → 'dd/mm/yyyy' או None (טקסט שאינו תאריך לא עולה לענן)."""
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    m = _DATE_RE.match(str(v).strip())
    if m:
        d, mo, y = m.groups()
        y = ("20" + y) if len(y) == 2 else y
        return f"{int(d):02d}/{int(mo):02d}/{y}"
    return None


def read_rows(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[SUMMARY_SHEET]
        rows = []
        for r in ws.iter_rows(min_row=2, max_col=17, values_only=True):
            name = str(r[1] or "").strip()
            if not name:
                continue
            streams = {}
            for col, label in STREAM_COLS:
                d = _fmt(r[col - 1])
                if d:
                    streams[label] = d
            rows.append({
                "referrer": str(r[0] or "").strip(),
                "name": name,
                "hp": str(r[4] or "").strip(),
                "streams": streams,
            })
        return rows
    finally:
        wb.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--api-base", default=DEFAULT_API_BASE)
    ap.add_argument("--masad-path", default=MASAD_PATH)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    token = os.environ.get("ECOOIL_BRIDGE_TOKEN")
    if not token and not args.dry_run:
        print("ECOOIL_BRIDGE_TOKEN missing"); return 1
    if not os.path.exists(args.masad_path):
        print(f"masad not found: {args.masad_path}"); return 1

    rows = read_rows(args.masad_path)
    with_ref = sum(1 for r in rows if r["referrer"])
    print(f"read {len(rows)} rows ({with_ref} with a referrer)")
    if args.dry_run:
        for r in rows[:5]:
            print(" ", json.dumps(r, ensure_ascii=False))
        return 0

    resp = requests.post(
        args.api_base + "/bridge/ecooil/validity-snapshot",
        headers={"Authorization": "Bearer " + token},
        json={"rows": rows}, timeout=60)
    print(f"push: {resp.status_code} {resp.text[:200]}")
    return 0 if resp.status_code == 200 else 1


if __name__ == "__main__":
    sys.exit(main())
