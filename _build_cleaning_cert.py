# -*- coding: utf-8 -*-
"""Build new-design cleaning certificate sheet inside טנקו_NEW_DESIGN.xlsx.
Also adds column AC ("קודי שטיפה") to all monthly log sheets.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

SRC = r'C:/Users/lio2e/Documents/AI_DEV/eco_oil_platform/טנקו_NEW_DESIGN.xlsx'
ASSETS = r'C:/Users/lio2e/Documents/AI_DEV/eco_oil_platform/website/images/cert'

wb = openpyxl.load_workbook(SRC)

# ===== Step 1: Add column AC ("קודי שטיפה") to each monthly log =====
MONTHLY = ['ינואר', 'פבואר ', 'מרץ', 'אפריל']
for sname in MONTHLY:
    if sname not in wb.sheetnames:
        continue
    ws_m = wb[sname]
    # row 2 holds column headers in this file
    ws_m.cell(row=2, column=29, value='קודי שטיפה')  # column AC = 29
    ws_m.column_dimensions['AC'].width = 28

# Sample data for the demo row 198 in אפריל so cert shows ✓ marks
if 'אפריל' in wb.sheetnames:
    wb['אפריל'].cell(row=198, column=29, value='P01, P09, C01, P30, T02, T20')

# ===== Step 2: Remove old cleaning cert sheet =====
for n in ['תעודת שטיפה ', 'תעודת שטיפה']:
    if n in wb.sheetnames:
        del wb[n]

ws = wb.create_sheet('תעודת שטיפה')
ws.sheet_view.rightToLeft = False

# ===== Column widths =====
COL_W = {'A': 3.5, 'B': 10, 'C': 12, 'D': 12, 'E': 3.5,
         'F': 3.5, 'G': 10, 'H': 12, 'I': 14, 'J': 3.5}
for col, w in COL_W.items():
    ws.column_dimensions[col].width = w

# ===== Styles =====
thin = Side(border_style='thin', color='000000')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

font_title   = Font(name='Arial', size=14, bold=True, color='000000')
font_certno  = Font(name='Arial', size=11, bold=True, color='000000')
font_sub     = Font(name='Arial', size=8,  italic=True, color='000000')
font_url     = Font(name='Arial', size=8,  color='666666')
font_idx     = Font(name='Arial', size=11, bold=True)
font_label   = Font(name='Arial', size=8,  bold=True, color='666666')
font_value   = Font(name='Arial', size=10)
font_value_b = Font(name='Arial', size=10, bold=True)
font_decl    = Font(name='Arial', size=9,  italic=True)
font_foot    = Font(name='Arial', size=8,  color='444444')
font_cap     = Font(name='Arial', size=9,  bold=True)

fill_black   = PatternFill('solid', fgColor='000000')
fill_lightg  = PatternFill('solid', fgColor='F0F0F0')

center     = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left',   vertical='center', wrap_text=True)
left_top   = Alignment(horizontal='left',   vertical='top',    wrap_text=True)


def merge(rng): ws.merge_cells(rng)


def set_cell(addr, value=None, font=None, fill=None, align=None, border=True):
    cell = ws[addr]
    if value is not None: cell.value = value
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border_all
    return cell


def setm(rng, value, font=None, fill=None, align=None, border=False):
    merge(rng)
    cell = ws[rng.split(':')[0]]
    cell.value = value
    if font:  cell.font = font
    if fill:  cell.fill = fill
    if align: cell.alignment = align
    if border:
        for r in ws[rng]:
            for c in (r if hasattr(r, '__iter__') else [r]):
                c.border = border_all


# ============ HEADER ============
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 22
ws.row_dimensions[3].height = 16
ws.row_dimensions[4].height = 14

# Left cell (A1:B4) for QR — wider so the QR is readable
merge('A1:B4')
ws['A1'].fill = PatternFill('solid', fgColor='FFFFFF')
ws['A1'].alignment = center
ws['A1'].border = border_all

setm('C1:H1', 'Cleaning Certificate', font_title, None, center, border=True)
setm('C2:H2', 'ECO-2026-0001', font_certno, None, center, border=True)
setm('C3:H3', 'Eco Oil Original Cleaning Certificate — please check validity online via QR code',
     font_sub, None, center, border=False)
setm('C4:H4', 'www.eco-oil.co.il', font_url, None, center, border=False)

# Right cell (I1:J4) for logo
merge('I1:J4')
ws['I1'].fill = PatternFill('solid', fgColor='FFFFFF')
ws['I1'].border = border_all

# ============ FORM TABLE ============
row = 5


def section(idx, label, value=None, value_height=22):
    global row
    ws.row_dimensions[row].height = 14
    if value is not None:
        ws.row_dimensions[row + 1].height = value_height
    set_cell(f'A{row}', value=str(idx), font=font_idx, fill=fill_lightg, align=center)
    if value is not None:
        merge(f'A{row}:A{row+1}')
    set_cell(f'B{row}', value=label, font=font_label, align=left_align)
    merge(f'B{row}:J{row}')
    if value is not None:
        set_cell(f'B{row+1}', value=value, font=font_value, align=left_top)
        merge(f'B{row+1}:J{row+1}')
        row += 2
    else:
        row += 1


def two_section(idx_l, lbl_l, val_l, idx_r, lbl_r, val_r, val_height=22):
    global row
    ws.row_dimensions[row].height = 14
    ws.row_dimensions[row + 1].height = val_height
    # left
    set_cell(f'A{row}', value=str(idx_l), font=font_idx, fill=fill_lightg, align=center)
    merge(f'A{row}:A{row+1}')
    set_cell(f'B{row}', value=lbl_l, font=font_label, align=left_align)
    merge(f'B{row}:E{row}')
    set_cell(f'B{row+1}', value=val_l, font=font_value, align=left_top)
    merge(f'B{row+1}:E{row+1}')
    # right
    set_cell(f'F{row}', value=str(idx_r), font=font_idx, fill=fill_lightg, align=center)
    merge(f'F{row}:F{row+1}')
    set_cell(f'G{row}', value=lbl_r, font=font_label, align=left_align)
    merge(f'G{row}:J{row}')
    set_cell(f'G{row+1}', value=val_r, font=font_value, align=left_top)
    merge(f'G{row+1}:J{row+1}')
    row += 2


# 1 - Cleaning station
station_text = ('Eco Oil Arrow Virometal Ltd.\n'
                '27 Haashlag St., Haifa 3660101\n'
                'P.O. Box 116, Nesher, 3660101\n'
                'P: +972-54-3232617\n'
                'Website: www.eco-oil.co.il')
section(1, 'Cleaning station', value=station_text, value_height=58)

# 2/3 - Customer ref / Internal ref
two_section(2, 'Customer ref. number', '=אפריל!G198',
            3, 'Internal ref.', '=אפריל!A198')

# 4/5 - Customer / Identification numbers
two_section(4, 'Customer', 'TANKO',
            5, 'Identification numbers',
            '="Equipment type: ISOTANK" & CHAR(10) & "Equipment N°: " & אפריל!A198',
            val_height=36)

# 6 - Nature of product
section(6, 'Nature of product', value='=אפריל!E198', value_height=22)

# 7 - Previous load
ws.row_dimensions[row].height = 14
set_cell(f'A{row}', value='7', font=font_idx, fill=fill_lightg, align=center)
merge(f'A{row}:A{row+2}')
set_cell(f'B{row}', value='Previous load', font=font_label, align=left_align)
merge(f'B{row}:J{row}')
row += 1
ws.row_dimensions[row].height = 14
for col_idx, header in [(2, 'Comp'), (3, 'UN N°'), (4, 'Name')]:
    align = center if col_idx != 4 else left_align
    set_cell(f'{openpyxl.utils.get_column_letter(col_idx)}{row}',
             value=header, font=font_label, fill=fill_lightg, align=align)
merge(f'D{row}:J{row}')
row += 1
ws.row_dimensions[row].height = 18
for col_idx in [2, 3, 4]:
    set_cell(f'{openpyxl.utils.get_column_letter(col_idx)}{row}',
             value='', font=font_value)
merge(f'D{row}:J{row}')
row += 1


# Helper for code/test rows: produces SEARCH-based ✓ formula
def chk_formula(code, source_addr='אפריל!AC198'):
    return f'=IF(ISNUMBER(SEARCH("{code}",{source_addr})),"✓","")'


def chk_t90():
    """T90 vacuum test pulls from column S of monthly log."""
    return '=IF(אפריל!S198>0,"✓","")'


def code_pair_row(left, right):
    """Render one row with two code cells: each is (chk_value, code, desc)."""
    global row
    ws.row_dimensions[row].height = 16
    if left:
        set_cell(f'B{row}', value=left[0], font=font_value_b, align=center)
        set_cell(f'C{row}', value=left[1], font=font_value_b, align=center)
        set_cell(f'D{row}', value=left[2], font=font_value, align=left_align)
        merge(f'D{row}:E{row}')
    else:
        for c in ['B','C','D','E']:
            set_cell(f'{c}{row}', value='', font=font_value)
        merge(f'D{row}:E{row}')
    if right:
        set_cell(f'F{row}', value=right[0], font=font_value_b, align=center)
        set_cell(f'G{row}', value=right[1], font=font_value_b, align=center)
        set_cell(f'H{row}', value=right[2], font=font_value, align=left_align)
        merge(f'H{row}:J{row}')
    else:
        for c in ['F','G','H','I','J']:
            set_cell(f'{c}{row}', value='', font=font_value)
        merge(f'H{row}:J{row}')
    row += 1


# 8 - Cleaning procedures (EFTCO codes) — 7 codes
cleaning_codes = [
    ('P01', 'Cold water spin'),
    ('P09', 'Hot water spin (T>80°C)'),
    ('C01', 'Alkaline detergent'),
    ('P40', 'Steaming'),
    ('C61', 'Acetone'),
    ('C60', 'Solvents'),
    ('P30', 'Drying'),
]
# Header row (label only)
ws.row_dimensions[row].height = 14
set_cell(f'A{row}', value='8', font=font_idx, fill=fill_lightg, align=center)
n_rows = (len(cleaning_codes) + 1) // 2  # rounded up
merge(f'A{row}:A{row+n_rows}')
set_cell(f'B{row}', value='Cleaning procedures (EFTCO codes)', font=font_label, align=left_align)
merge(f'B{row}:J{row}')
row += 1
# Pair them up: 2 columns
for i in range(0, len(cleaning_codes), 2):
    left  = cleaning_codes[i]
    right = cleaning_codes[i+1] if i+1 < len(cleaning_codes) else None
    left_tup  = (chk_formula(left[0]),  left[0],  left[1])  if left  else None
    right_tup = (chk_formula(right[0]), right[0], right[1]) if right else None
    code_pair_row(left_tup, right_tup)

# 9 - Tests — 5 codes (T01, T02, T20, T50, T90)
test_codes = [
    ('T01', 'Visual / Odour inspection', chk_formula('T01')),
    ('T02', 'Visual inspection',          chk_formula('T02')),
    ('T20', 'pH neutral',                 chk_formula('T20')),
    ('T50', 'Air test',                   chk_formula('T50')),
    ('T90', 'Vacuum test',                chk_t90()),
]
ws.row_dimensions[row].height = 14
set_cell(f'A{row}', value='9', font=font_idx, fill=fill_lightg, align=center)
n_test_rows = (len(test_codes) + 1) // 2
merge(f'A{row}:A{row+n_test_rows}')
set_cell(f'B{row}', value='Tests', font=font_label, align=left_align)
merge(f'B{row}:J{row}')
row += 1
for i in range(0, len(test_codes), 2):
    left  = test_codes[i]
    right = test_codes[i+1] if i+1 < len(test_codes) else None
    left_tup  = (left[2],  left[0],  left[1])  if left  else None
    right_tup = (right[2], right[0], right[1]) if right else None
    code_pair_row(left_tup, right_tup)

# 10 - Additional services (8 items, but pull formulas from log where possible)
services = [
    ('Transportation', '=IF(אפריל!U198>0,"✓","")', 'Photo set',   '=IF(אפריל!T198>0,"✓","")'),
    ('Polish',         '',                          'Vacuum test', '=IF(אפריל!S198>0,"✓","")'),
    ('Repair',         '=IF(אפריל!V198>0,"✓","")', 'Storage',     '=IF(אפריל!P198>0,"✓","")'),
    ('Test',           '',                          'Maintenance', ''),
]
ws.row_dimensions[row].height = 14
set_cell(f'A{row}', value='10', font=font_idx, fill=fill_lightg, align=center)
merge(f'A{row}:A{row+len(services)}')
set_cell(f'B{row}', value='Additional services', font=font_label, align=left_align)
merge(f'B{row}:J{row}')
row += 1
for s in services:
    ws.row_dimensions[row].height = 16
    set_cell(f'B{row}', value=s[0], font=font_value, align=left_align)
    merge(f'B{row}:D{row}')
    set_cell(f'E{row}', value=s[1], font=font_value_b, align=center)
    set_cell(f'F{row}', value=s[2], font=font_value, align=left_align)
    merge(f'F{row}:I{row}')
    set_cell(f'J{row}', value=s[3], font=font_value_b, align=center)
    row += 1

# 11 - Comments
section(11, 'Comments', value='=אפריל!H198', value_height=28)

# 12 - Date Time in / End of Cleaning
two_section(12, 'Date, Time in',
            '=TEXT(אפריל!D198,"dd/mm/yyyy")&" "&TEXT(אפריל!M198,"hh:mm")',
            '', 'End of Cleaning — Date, Time out',
            '=TEXT(אפריל!L198,"dd/mm/yyyy")&" "&TEXT(אפריל!M198,"hh:mm")',
            val_height=22)

# Declaration
ws.row_dimensions[row].height = 22
decl = ('The cleaning station and the driver confirm that the above '
        'service(s) to clean the tank have been carried out.')
set_cell(f'A{row}', value=decl, font=font_decl, align=center)
merge(f'A{row}:J{row}')
row += 1

# 13 - Signature + stamp
ws.row_dimensions[row].height = 14
set_cell(f'A{row}', value='13', font=font_idx, fill=fill_lightg, align=center)
merge(f'A{row}:A{row+4}')
set_cell(f'B{row}', value='Cleaning station', font=font_label, align=left_align)
merge(f'B{row}:J{row}')
row += 1

sig_row = row
ws.row_dimensions[row].height = 50
ws.row_dimensions[row + 1].height = 50
set_cell(f'B{row}', value='', align=center)
merge(f'B{row}:E{row+1}')
set_cell(f'F{row}', value='', align=center)
merge(f'F{row}:J{row+1}')
row += 2

ws.row_dimensions[row].height = 4
set_cell(f'B{row}', value='')
merge(f'B{row}:E{row}')
set_cell(f'F{row}', value='')
merge(f'F{row}:J{row}')
row += 1

ws.row_dimensions[row].height = 18
set_cell(f'B{row}', value='CEO Signature – Eco Oil', font=font_cap, align=center)
merge(f'B{row}:E{row}')
set_cell(f'F{row}', value='Company Stamp', font=font_cap, align=center)
merge(f'F{row}:J{row}')
row += 1

# Footer image row
ws.row_dimensions[row].height = 130
set_cell(f'A{row}', value='', align=center)
merge(f'A{row}:J{row}')
foot_row = row
row += 1

# Print info
ws.row_dimensions[row].height = 14
set_cell(f'A{row}', value='Printed with status SIGNED by Eco Oil — www.eco-oil.co.il',
         font=font_foot, align=center)
merge(f'A{row}:J{row}')
last_row = row

# ===== Insert images =====
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image as PILImage

EMU = 9525  # EMU per pixel


def add_centered_image(img_path, anchor_col_idx, anchor_row_idx,
                       max_width_px, max_height_px,
                       area_width_px, area_height_px):
    """Insert image centered within an area, preserving aspect ratio."""
    pil = PILImage.open(img_path)
    orig_w, orig_h = pil.size
    ratio = orig_w / orig_h
    # fit within max box while preserving ratio
    if max_width_px / max_height_px > ratio:
        h = max_height_px
        w = int(h * ratio)
    else:
        w = max_width_px
        h = int(w / ratio)
    # center within area
    off_x = max(0, (area_width_px - w) // 2)
    off_y = max(0, (area_height_px - h) // 2)
    img = XLImage(img_path)
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=anchor_col_idx, row=anchor_row_idx,
                           colOff=off_x * EMU, rowOff=off_y * EMU),
        ext=XDRPositiveSize2D(cx=w * EMU, cy=h * EMU),
    )
    ws._images.append(img)


# QR (top-left of header) - anchor to A1 (col=0, row=0)
# A:B cols widths 3.5 + 10 = 13.5 char ≈ 100 px
add_centered_image(f'{ASSETS}/qr-eco-oil.png',
                   anchor_col_idx=0, anchor_row_idx=0,
                   max_width_px=78, max_height_px=78,
                   area_width_px=100, area_height_px=88)

# Logo (top-right of header) - anchor to I1 (col=8, row=0)
# I:J cols widths 14 + 3.5 = 17.5 char ≈ 125 px
add_centered_image(f'{ASSETS}/eco-oil-logo.png',
                   anchor_col_idx=8, anchor_row_idx=0,
                   max_width_px=80, max_height_px=80,
                   area_width_px=125, area_height_px=88)

# Signature merged B..E across 2 rows. B = col 1 (0-based)
# Approximate area: cols B(10) C(12) D(12) E(3.5) ≈ 37.5 char × ~7 px ≈ 260 px
# Rows: 50 + 50 = 100 px
add_centered_image(f'{ASSETS}/signature-yoav.png',
                   anchor_col_idx=1, anchor_row_idx=sig_row - 1,
                   max_width_px=120, max_height_px=92,
                   area_width_px=260, area_height_px=100)

# Stamp merged F..J across 2 rows. F = col 5 (0-based)
# Cols F(3.5) G(10) H(12) I(14) J(3.5) ≈ 43 char × ~7 px ≈ 300 px
add_centered_image(f'{ASSETS}/stamp.png',
                   anchor_col_idx=5, anchor_row_idx=sig_row - 1,
                   max_width_px=240, max_height_px=92,
                   area_width_px=300, area_height_px=100)

# Footer contact strip (full width A..J)
add_centered_image(f'{ASSETS}/footer-contact.png',
                   anchor_col_idx=0, anchor_row_idx=foot_row - 1,
                   max_width_px=595, max_height_px=128,
                   area_width_px=600, area_height_px=130)

# ===== Page setup =====
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.4
ws.print_area = f'A1:J{last_row}'
ws.print_options.horizontalCentered = True

# Reorder sheets
target = ['ינואר', 'פבואר ', 'מרץ', 'אפריל', 'תעודת שטיפה', 'תעודת יציאה ']
order_idx = {n: i for i, n in enumerate(target)}
wb._sheets.sort(key=lambda s: order_idx.get(s.title, 99))

wb.save(SRC)
print('Saved:', SRC)
print('Last row:', last_row)
