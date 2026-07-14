# -*- coding: utf-8 -*-
"""Build a focused worklist of on-site isotanks that are MISSING data, for Limor +
the worker to complete one-by-one (with the live-file ROW NUMBER for fast lookup).
Only incomplete tanks are listed — fully-complete ones are excluded on purpose."""
import os
os.environ.pop("DATABASE_URL", None)
from src.app import create_app
from src.app.db import DepotIsotankVisit
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

ON_SITE = ["באחסון", "בטיפול שטיפה", "בתיקון", "מוכן לשחרור", "הכנה לשחרור"]
OUT = r"C:\eco_oil_portal\השלמת נתונים - מכלים באתר.xlsx"

def fill(c): return PatternFill(fill_type="solid", fgColor=c)
brand=fill("5B9E96"); dark=fill("3E6F69"); grp=fill("D6E4E1"); miss=fill("FCE4D6")
white=Font(bold=True,color="FFFFFF"); white_l=Font(bold=True,color="FFFFFF",size=14)
bold=Font(bold=True); thin=Side(style="thin",color="C9C9C9")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
rt=Alignment(horizontal="right",vertical="center")

def has(v): return v not in (None, "", "—")
def fmtd(d): return d.strftime("%d/%m/%Y") if d else None
def incomplete(v): return (not v.storage_in_date) or (not has(v.last_material))

app = create_app()
with app.app_context():
    allrows = DepotIsotankVisit.query.filter(DepotIsotankVisit.status.in_(ON_SITE)).all()
    rows = [v for v in allrows if incomplete(v)]
    by_company = {}
    for v in rows:
        by_company.setdefault(v.company or "(ללא לקוח)", []).append(v)
    order = sorted(by_company, key=lambda c: -len(by_company[c]))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "מכלים להשלמה"
ws.sheet_view.rightToLeft = True; ws.sheet_view.showGridLines = False

ws.merge_cells("A1:G1")
c = ws.cell(1,1,"השלמת נתונים — מכלים באתר שחסר בהם מידע   (לפי הקובץ החי)"); c.fill=brand; c.font=white_l; c.alignment=ctr
ws.row_dimensions[1].height = 30

headers = ["שורה בקובץ", "מספר מכל", "סטטוס", "תאריך כניסה לאחסון", "חומר אחרון", "מה חסר", "טופל ✓"]
for j,h in enumerate(headers, start=1):
    cc = ws.cell(2,j,h); cc.fill=dark; cc.font=white; cc.alignment=ctr; cc.border=border
ws.row_dimensions[2].height = 24

r = 3
for comp in order:
    items = sorted(by_company[comp], key=lambda v: (v.source_row or 0))
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
    gc = ws.cell(r,1, f"לקוח: {comp}   —   {len(items)} מכלים להשלמה")
    gc.fill=grp; gc.font=bold; gc.alignment=rt
    ws.row_dimensions[r].height = 22; r += 1
    for v in items:
        d_txt = fmtd(v.storage_in_date) or "חסר"
        m_txt = v.last_material if has(v.last_material) else "חסר"
        missing = []
        if not v.storage_in_date: missing.append("תאריך אחסון")
        if not has(v.last_material): missing.append("חומר")
        vals = [v.source_row, v.isotank_number, v.status, d_txt, m_txt, " + ".join(missing), ""]
        for j,val in enumerate(vals, start=1):
            cell = ws.cell(r,j,val); cell.alignment=ctr; cell.border=border
            if j==1: cell.font=bold
        if d_txt=="חסר": ws.cell(r,4).fill=miss; ws.cell(r,4).font=bold
        if m_txt=="חסר": ws.cell(r,5).fill=miss; ws.cell(r,5).font=bold
        ws.cell(r,6).fill = miss
        r += 1

for col,w in zip("ABCDEFG",[11,16,14,18,22,18,9]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A3"
wb.save(OUT)
print(f"WROTE {OUT} | tanks to complete={len(rows)} (of {len(allrows)} on-site) | companies={len(order)}")
